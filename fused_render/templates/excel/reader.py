# /// script
# dependencies = ["fpdf2>=2.8.7"]
# ///
"""Data ops for the excel editor view.

Two data paths:

* **Rich path** (xlsx where every sheet is small): full load/save through
  openpyxl with formulas and cell formatting — the original behavior.
* **DuckDB path** (big sheets, or .csv/.parquet sources): each sheet is
  converted once into an all-VARCHAR Parquet cache (columns c0..cN plus a
  stable __rid row id) under ~/.fused-render/cache/excel/, and the browser
  fetches windowed row batches. Sort and filter run as SQL over the whole
  dataset, so nothing is ever truncated. Edits travel back as a sparse
  {"rid:col": value} overlay and are merged on save. No data loss at any size.

`file` is always the workbook the template was opened for (the read-only
`_file` param) — this editor edits one file in place; it does not browse to
or open a different one. Save writes back to `file`; save_as/export write
wherever the caller points them.

Actions (dispatched via the `action` param):
  load        — sheet metadata + inline rows (small) or first row batch (big)
  rows        — windowed batch: offset/limit + server-side sort/filter
  save        — write back: rich xlsx / streamed xlsx / csv / parquet
  save_as     — write to a user-chosen directory + name, return the new path
  export      — write csv / xlsx / pdf / parquet into the exports cache dir
  run_script  — exec a user Python script against (small-)sheet data
  listdir     — directory listing for the save-as folder browser
"""

import json
import os
import re

# The fused engine execs this script without setting __file__; it puts the
# script's own directory first on sys.path, so rebuild __file__ from it. Under
# the built-in executor __file__ is already set, so this is a no-op.
if "__file__" not in globals():
    import os, sys
    __file__ = os.path.join(sys.path[0], "reader.py")

HERE = os.path.dirname(os.path.abspath(__file__))

# Everything derived from a workbook (DuckDB parquet cache, export scratch
# copies) lives under the user's cache dir, never next to the template.
CACHE_ROOT = os.path.expanduser(os.path.join("~", ".fused-render", "cache", "excel"))
EXPORTS = os.path.join(CACHE_ROOT, "exports")

SMALL_ROWS = 10_000        # a sheet bigger than this in either measure ...
SMALL_CELLS = 200_000      # ... goes to the DuckDB path instead of inline JSON
FIRST_BATCH = 500
XLSX_SAVE_MAX_ROWS = 150_000  # streamed xlsx writes above this exceed the 30 s call budget

SPREADSHEET_EXTS = (".xlsx", ".xlsm", ".csv", ".parquet")


def _safe_name(name, default):
    name = re.sub(r"[^\w.\- ()]+", "_", os.path.basename(str(name or default)).strip())
    return name or default


# ---------- cell value conversion ----------

def _cell_out(v):
    if v is None:
        return ""
    if isinstance(v, (int, float, bool, str)):
        return v
    return str(v)


def _cell_in(s):
    if not isinstance(s, str):
        return s
    t = s.strip()
    if t == "":
        return None
    if t.startswith("="):
        return t
    low = t.lower()
    if low in ("true", "false"):
        return low == "true"
    try:
        return int(t)
    except ValueError:
        pass
    try:
        return float(t)
    except ValueError:
        pass
    return s


# ---------- style conversion (rich path only) ----------

_FMT_TO_EXCEL = {"number": "#,##0.00", "percent": "0.00%", "currency": '"$"#,##0.00'}


def _style_out(cell):
    st = {}
    f = cell.font
    if f is not None:
        if f.b:
            st["b"] = 1
        if f.i:
            st["i"] = 1
        if f.u:
            st["u"] = 1
        if f.color is not None and f.color.type == "rgb" and isinstance(f.color.rgb, str):
            hexpart = f.color.rgb[-6:].upper()
            if hexpart not in ("000000",):
                st["color"] = "#" + hexpart
    fill = cell.fill
    if fill is not None and fill.fill_type == "solid" and fill.fgColor is not None:
        if fill.fgColor.type == "rgb" and isinstance(fill.fgColor.rgb, str):
            hexpart = fill.fgColor.rgb[-6:].upper()
            if fill.fgColor.rgb not in ("00000000",):
                st["bg"] = "#" + hexpart
    al = cell.alignment
    if al is not None and al.horizontal in ("left", "center", "right"):
        st["align"] = al.horizontal
    nf = cell.number_format or ""
    if nf != "General":
        if "%" in nf:
            st["fmt"] = "percent"
        elif "$" in nf:
            st["fmt"] = "currency"
        elif "#,##0" in nf or nf.startswith("0.0"):
            st["fmt"] = "number"
    return st or None


def _style_kwargs(st):
    from openpyxl.styles import Alignment, Font, PatternFill

    font_kw = {}
    if st.get("b"):
        font_kw["bold"] = True
    if st.get("i"):
        font_kw["italic"] = True
    if st.get("u"):
        font_kw["underline"] = "single"
    if st.get("color"):
        font_kw["color"] = "FF" + st["color"].lstrip("#").upper()
    out = {}
    if font_kw:
        out["font"] = Font(**font_kw)
    if st.get("bg"):
        out["fill"] = PatternFill("solid", fgColor="FF" + st["bg"].lstrip("#").upper())
    if st.get("align"):
        out["alignment"] = Alignment(horizontal=st["align"])
    if st.get("fmt") in _FMT_TO_EXCEL:
        out["number_format"] = _FMT_TO_EXCEL[st["fmt"]]
    return out


def _style_in(cell, st):
    if not st:
        return
    for attr, val in _style_kwargs(st).items():
        setattr(cell, attr, val)


# ---------- duckdb helpers ----------

def _duck(excel_ext=False):
    import duckdb

    con = duckdb.connect()
    if excel_ext:
        con.execute("INSTALL excel; LOAD excel;")
    return con


def _q(s):
    return "'" + str(s).replace("'", "''") + "'"


def _copy_to_parquet(con, src_sql, pq):
    """Normalize any relation to (__rid BIGINT, c0..cN VARCHAR) parquet."""
    cols = [d[0] for d in con.execute(f"SELECT * FROM ({src_sql}) LIMIT 0").description]
    sel = ", ".join(f'CAST("{c}" AS VARCHAR) AS c{i}' for i, c in enumerate(cols))
    con.execute(
        f"COPY (SELECT row_number() OVER () - 1 AS __rid, {sel} FROM ({src_sql})) "
        f"TO {_q(pq)} (FORMAT PARQUET)"
    )
    n = con.execute(f"SELECT count(*) FROM read_parquet({_q(pq)})").fetchone()[0]
    return n, len(cols), cols


def _xlsx_sheet_to_parquet(file, sheet, pq):
    try:
        con = _duck(excel_ext=True)
        n, nc, _ = _copy_to_parquet(
            con,
            f"SELECT * FROM read_xlsx({_q(file)}, sheet={_q(sheet)}, header=false, all_varchar=true)",
            pq,
        )
        return n, nc
    except Exception:
        # fallback: stream with openpyxl (slower, fine for medium files)
        import openpyxl
        import pyarrow as pa
        import pyarrow.parquet as papq

        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb[sheet]
        data, ncols = [], 0
        for row in ws.iter_rows(values_only=True):
            r = ["" if v is None else str(v) for v in row]
            ncols = max(ncols, len(r))
            data.append(r)
        wb.close()
        ncols = max(ncols, 1)
        cols = {"__rid": list(range(len(data)))}
        for c in range(ncols):
            cols[f"c{c}"] = [r[c] if c < len(r) else "" for r in data]
        papq.write_table(pa.table(cols), pq)
        return len(data), ncols


def _xlsx_dims(file):
    import openpyxl

    wb = openpyxl.load_workbook(file, read_only=True)
    dims = [(name, wb[name].max_row or 1, wb[name].max_column or 1) for name in wb.sheetnames]
    wb.close()
    return dims


def _is_big(nrows, ncols):
    return nrows > SMALL_ROWS or nrows * max(ncols, 1) > SMALL_CELLS


# ---------- parquet cache ----------

def _cache_dir(file):
    import hashlib

    h = hashlib.sha1(os.path.abspath(file).encode()).hexdigest()[:16]
    return os.path.join(CACHE_ROOT, f"{h}-{int(os.path.getmtime(file) * 1000)}")


def _clean_stale(keep_dir):
    if not os.path.isdir(CACHE_ROOT):
        return
    prefix = os.path.basename(keep_dir).split("-")[0]
    for n in os.listdir(CACHE_ROOT):
        p = os.path.join(CACHE_ROOT, n)
        if n.startswith(prefix + "-") and p != keep_dir:
            import shutil

            shutil.rmtree(p, ignore_errors=True)


def _ensure_cache(file):
    """Build (or reuse) the parquet cache + metadata for a workbook."""
    file = os.path.abspath(file)
    d = _cache_dir(file)
    meta_path = os.path.join(d, "meta.json")
    if os.path.exists(meta_path):
        with open(meta_path) as f:
            return json.load(f)
    _clean_stale(d)
    os.makedirs(d, exist_ok=True)
    ext = os.path.splitext(file)[1].lower()
    sheets = []
    if ext in (".xlsx", ".xlsm"):
        for i, (name, nr, nc) in enumerate(_xlsx_dims(file)):
            entry = {"name": name, "kind": "xlsx", "big": _is_big(nr, nc),
                     "nrows": nr, "ncols": nc, "header": None}
            if entry["big"]:
                pq = os.path.join(d, f"s{i}.parquet")
                nr2, nc2 = _xlsx_sheet_to_parquet(file, name, pq)
                entry.update(nrows=nr2, ncols=nc2, parquet=os.path.basename(pq))
            sheets.append(entry)
    elif ext == ".csv":
        con = _duck()
        pq = os.path.join(d, "s0.parquet")
        nr, nc, _ = _copy_to_parquet(
            con, f"SELECT * FROM read_csv({_q(file)}, header=false, all_varchar=true, null_padding=true)", pq
        )
        sheets.append({"name": os.path.splitext(os.path.basename(file))[0], "kind": "csv",
                       "big": _is_big(nr, nc), "nrows": nr, "ncols": nc, "header": None,
                       "parquet": "s0.parquet"})
    elif ext == ".parquet":
        con = _duck()
        pq = os.path.join(d, "s0.parquet")
        nr, nc, cols = _copy_to_parquet(con, f"SELECT * FROM read_parquet({_q(file)})", pq)
        sheets.append({"name": os.path.splitext(os.path.basename(file))[0], "kind": "parquet",
                       "big": _is_big(nr, nc), "nrows": nr, "ncols": nc, "header": cols,
                       "parquet": "s0.parquet"})
    else:
        raise ValueError(f"unsupported file type {ext!r}")
    meta = {"file": file, "mtime": os.path.getmtime(file), "sheets": sheets}
    with open(meta_path, "w") as f:
        json.dump(meta, f)
    return meta


def _rekey_cache(file):
    """After a save changed the file's mtime, move the cache to the new key."""
    import shutil

    file = os.path.abspath(file)
    new_dir = _cache_dir(file)
    if os.path.isdir(new_dir):
        return
    prefix = os.path.basename(new_dir).split("-")[0]
    if os.path.isdir(CACHE_ROOT):
        for n in sorted(os.listdir(CACHE_ROOT), reverse=True):
            if n.startswith(prefix + "-"):
                shutil.move(os.path.join(CACHE_ROOT, n), new_dir)
                meta_path = os.path.join(new_dir, "meta.json")
                try:
                    with open(meta_path) as f:
                        meta = json.load(f)
                    meta["mtime"] = os.path.getmtime(file)
                    with open(meta_path, "w") as f:
                        json.dump(meta, f)
                except OSError:
                    pass
                return


def _sheet_meta(meta, sheet_name):
    for i, sh in enumerate(meta["sheets"]):
        if sh["name"] == sheet_name:
            return i, sh
    return 0, meta["sheets"][0]


def _sheet_parquet(file, sh):
    return os.path.join(_cache_dir(file), sh["parquet"])


# ---------- windowed queries ----------

def _criteria_sql(col, q):
    """(sql, params) for a single text/number criteria like '>5' or 'foo'."""
    m = re.match(r"^(<>|<=|>=|=|<|>)(.*)$", q)
    if m and m.group(2).strip() != "":
        op, val = m.group(1), m.group(2).strip()
        try:
            num = float(val)
            if op == "=":
                return f"TRY_CAST({col} AS DOUBLE) = ?", [num]
            if op == "<>":
                return f"TRY_CAST({col} AS DOUBLE) IS DISTINCT FROM ?", [num]
            return f"TRY_CAST({col} AS DOUBLE) {op} ?", [num]
        except ValueError:
            sql_op = {"=": "=", "<>": "!=", "<": "<", ">": ">", "<=": "<=", ">=": ">="}[op]
            return f"lower(coalesce({col}, '')) {sql_op} lower(?)", [val]
    esc = q.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
    return f"{col} ILIKE ? ESCAPE '\\'", [f"%{esc}%"]


def _filters_sql(filters):
    """Build a WHERE for a list of {c, q?, exclude?} column filters. `q` is the
    optional condition string; `exclude` is a list of display values to hide
    (the Google-Sheets value checklist — everything not in the visible set)."""
    where, params = [], []
    for f in filters or []:
        c = int(f["c"])
        col = f"c{c}"
        q = str(f.get("q", "")).strip()
        if q:
            sql, ps = _criteria_sql(col, q)
            where.append(sql)
            params.extend(ps)
        exclude = f.get("exclude") or []
        if exclude:
            ph = ", ".join(["?"] * len(exclude))
            where.append(f"coalesce(CAST({col} AS VARCHAR), '') NOT IN ({ph})")
            params.extend(str(v) for v in exclude)
    return (" WHERE " + " AND ".join(where)) if where else "", params


def _query_distinct(pq, col, filters, cap=2000):
    """Distinct display values of one column, honouring the OTHER columns'
    filters (matching how Google Sheets narrows the value list). Returns up to
    `cap` values sorted, plus a truncated flag."""
    con = _duck()
    con.execute(f"CREATE VIEW t AS SELECT * FROM read_parquet({_q(pq)})")
    where, params = _filters_sql([f for f in (filters or []) if int(f["c"]) != col])
    rows = con.execute(
        f"SELECT DISTINCT coalesce(CAST(c{col} AS VARCHAR), '') AS v FROM t{where} "
        f"ORDER BY v LIMIT ?",
        params + [cap + 1],
    ).fetchall()
    vals = [r[0] for r in rows]
    return {"values": vals[:cap], "truncated": len(vals) > cap}


def _query_rows(pq, ncols, offset, limit, sort, filters):
    con = _duck()
    con.execute(f"CREATE VIEW t AS SELECT * FROM read_parquet({_q(pq)})")
    where, params = _filters_sql(filters)
    matched = con.execute(f"SELECT count(*) FROM t{where}", params).fetchone()[0]
    order = "__rid"
    if sort and sort.get("c") is not None and int(sort["c"]) >= 0:
        c = int(sort["c"])
        sd = "DESC" if sort.get("dir") == "desc" else "ASC"
        order = f"TRY_CAST(c{c} AS DOUBLE) {sd} NULLS LAST, c{c} {sd} NULLS LAST"
    sel = ", ".join(["__rid"] + [f"c{i}" for i in range(ncols)])
    rows = con.execute(
        f"SELECT {sel} FROM t{where} ORDER BY {order} LIMIT ? OFFSET ?",
        params + [int(limit), int(offset)],
    ).fetchall()
    rows = [[r[0]] + ["" if v is None else v for v in r[1:]] for r in rows]
    return {"rows": rows, "matched": matched}


def _table_with_edits(con, pq, edits, ncols):
    """Materialize the cached parquet with a sparse edits overlay applied."""
    con.execute(f"CREATE TABLE t AS SELECT * FROM read_parquet({_q(pq)})")
    for key, v in (edits or {}).items():
        rid_s, c_s = str(key).split(":")
        c = int(c_s)
        if 0 <= c < ncols:
            con.execute(f"UPDATE t SET c{c} = ? WHERE __rid = ?", [str(v), int(rid_s)])
    return "t"


# ---------- load ----------

def _ro_verdict(file):
    """Editability verdict for the viewed file (SPEC §13.5 RO-4).

    Folded into every load payload so the UI can badge read-only files and
    disable in-place saving while keeping viewing/export/save-as working.
    """
    if os.path.exists(file) and not os.access(file, os.W_OK):
        return {
            "editable": False,
            "readonly_message": "Read-only",
            "readonly_tooltip": "The file is read-only — its permissions "
                                "don't allow writing, so it can't be edited here.",
        }
    return {"editable": True, "readonly_message": "", "readonly_tooltip": ""}


def _sheet_view_out(ws):
    """Column widths (px) and frozen-row count from a worksheet."""
    from openpyxl.utils import column_index_from_string

    colw = {}
    for letter, dim in ws.column_dimensions.items():
        if dim.width:
            colw[str(column_index_from_string(letter) - 1)] = int(dim.width * 7 + 5)
    freeze = 0
    if ws.freeze_panes:
        m = re.match(r"[A-Z]+(\d+)", str(ws.freeze_panes))
        if m:
            freeze = int(m.group(1)) - 1
    return colw, freeze


def _sheet_view_in(ws, sh):
    from openpyxl.utils import get_column_letter

    for c_str, px in (sh.get("colw") or {}).items():
        try:
            ws.column_dimensions[get_column_letter(int(c_str) + 1)].width = max(1, round((float(px) - 5) / 7, 2))
        except (ValueError, TypeError):
            continue
    if sh.get("freeze"):
        ws.freeze_panes = f"A{int(sh['freeze']) + 1}"


def _load_rich(file):
    """Original openpyxl path: formulas + styles, all sheets small."""
    import openpyxl

    wb = openpyxl.load_workbook(file, data_only=False)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows, styles = [], []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1):
            rows.append([_cell_out(c.value) for c in row])
            styles.append([_style_out(c) for c in row])
        colw, freeze = _sheet_view_out(ws)
        sheets.append({"name": name, "rows": rows, "styles": styles, "big": False,
                       "colw": colw, "freeze": freeze})
    return {"sheets": sheets, "mtime": os.path.getmtime(file), "rich": True,
            **_ro_verdict(file)}


def _load(file):
    file = os.path.abspath(file)
    ext = os.path.splitext(file)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        if not any(_is_big(nr, nc) for _, nr, nc in _xlsx_dims(file)):
            return _load_rich(file)
    meta = _ensure_cache(file)
    d = _cache_dir(file)
    sheets = []
    for i, sh in enumerate(meta["sheets"]):
        if sh["big"]:
            first = _query_rows(os.path.join(d, sh["parquet"]), sh["ncols"], 0, FIRST_BATCH, None, None)
            sheets.append({"name": sh["name"], "big": True, "kind": sh["kind"],
                           "nrows": sh["nrows"], "ncols": sh["ncols"], "header": sh["header"],
                           "rows": first["rows"], "matched": first["matched"]})
        elif sh["kind"] == "xlsx":
            # small sheet inside a workbook that also has big ones: stream values
            import openpyxl

            wb = openpyxl.load_workbook(file, read_only=True, data_only=False)
            ws = wb[sh["name"]]
            rows = [[_cell_out(v) for v in row] for row in ws.iter_rows(values_only=True)] or [[""]]
            wb.close()
            sheets.append({"name": sh["name"], "rows": rows, "styles": None, "big": False, "kind": "xlsx"})
        else:
            con = _duck()
            sel = ", ".join(f"c{i2}" for i2 in range(sh["ncols"]))
            data = con.execute(
                f"SELECT {sel} FROM read_parquet({_q(os.path.join(d, sh['parquet']))}) ORDER BY __rid"
            ).fetchall()
            rows = [["" if v is None else v for v in r] for r in data] or [[""]]
            sheets.append({"name": sh["name"], "rows": rows, "styles": None, "big": False,
                           "kind": sh["kind"], "header": sh["header"]})
    return {"sheets": sheets, "mtime": os.path.getmtime(file), "rich": False,
            **_ro_verdict(file)}


# ---------- save ----------

def _build_workbook(sheets):
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sh in sheets:
        ws = wb.create_sheet(str(sh["name"])[:31] or "Sheet")
        _sheet_view_in(ws, sh)
        styles = sh.get("styles") or []
        for r, row in enumerate(sh["rows"], start=1):
            srow = styles[r - 1] if r - 1 < len(styles) else []
            for c, v in enumerate(row, start=1):
                val = _cell_in(v)
                st = srow[c - 1] if c - 1 < len(srow) else None
                if val is None and not st:
                    continue
                cell = ws.cell(row=r, column=c)
                if val is not None:
                    cell.value = val
                _style_in(cell, st)
    return wb


def _persist_edits(file, sheets_payload, meta):
    """Merge each big sheet's edits overlay into its cache parquet (before the
    file itself is rewritten — the cache key includes the file's mtime)."""
    for sp in sheets_payload:
        if sp.get("big") and sp.get("edits"):
            _, sh = _sheet_meta(meta, sp["name"])
            pq = _sheet_parquet(file, sh)
            con = _duck()
            _table_with_edits(con, pq, sp["edits"], sh["ncols"])  # materializes in memory
            con.execute(f"COPY (SELECT * FROM t ORDER BY __rid) TO {_q(pq)} (FORMAT PARQUET)")


def _write_xlsx_streamed(dest, file, sheets_payload, meta):
    """write_only workbook to `dest`; small sheets come from inline rows, big
    sheets stream from `file`'s (already edit-merged) cache parquet."""
    import openpyxl
    from openpyxl.cell import WriteOnlyCell

    for sp in sheets_payload:
        if sp.get("big"):
            _, sh = _sheet_meta(meta, sp["name"])
            if sh["nrows"] > XLSX_SAVE_MAX_ROWS:
                raise ValueError(
                    f"Sheet “{sp['name']}” has {sh['nrows']:,} rows — too large to write back "
                    f"as .xlsx within the call budget. Use File ▸ Download as CSV/Parquet, or "
                    f"open the data as .csv/.parquet (saves at any size)."
                )
    wb = openpyxl.Workbook(write_only=True)
    for sp in sheets_payload:
        ws = wb.create_sheet(str(sp["name"])[:31] or "Sheet")
        _sheet_view_in(ws, sp)  # write_only allows dims/freeze before rows
        if sp.get("big"):
            _, sh = _sheet_meta(meta, sp["name"])
            con = _duck()
            sel = ", ".join(f"c{i}" for i in range(sh["ncols"]))
            cur = con.execute(
                f"SELECT {sel} FROM read_parquet({_q(_sheet_parquet(file, sh))}) ORDER BY __rid"
            )
            while True:
                chunk = cur.fetchmany(10_000)
                if not chunk:
                    break
                for row in chunk:
                    ws.append([_cell_in(v) if v is not None else None for v in row])
        else:
            styles = sp.get("styles") or []
            for r, row in enumerate(sp["rows"]):
                srow = styles[r] if r < len(styles) else []
                out = []
                for c, v in enumerate(row):
                    val = _cell_in(v)
                    st = srow[c] if c < len(srow) else None
                    if st:
                        cell = WriteOnlyCell(ws, value=val)
                        for attr, sv in _style_kwargs(st).items():
                            setattr(cell, attr, sv)
                        out.append(cell)
                    else:
                        out.append(val)
                out.append(None)  # openpyxl drops trailing Nones; harmless
                ws.append(out[:-1])
    wb.save(dest)


def _save(file, sheets_payload, expected_mtime):
    file = os.path.abspath(file)
    # SPEC §13.5 RO-3: every branch below writes via tmp + os.replace, which
    # goes through the (writable) parent dir and would silently bypass a
    # chmod -w bit on the file itself — gate explicitly, before the conflict
    # check (offering "Overwrite" on an unwritable file would be a lie).
    # Existence-qualified: _save_as points here at a fresh dest, and
    # os.access on a nonexistent path is False.
    if os.path.exists(file) and not os.access(file, os.W_OK):
        raise PermissionError(f"{file!r} is read-only")
    if expected_mtime and os.path.exists(file):
        on_disk = os.path.getmtime(file)
        if abs(on_disk - float(expected_mtime)) > 1e-6:
            return {"conflict": True, "mtime": on_disk}
    ext = os.path.splitext(file)[1].lower()
    any_big = any(s.get("big") for s in sheets_payload)
    tmp = file + ".tmp"

    if ext in (".xlsx", ".xlsm"):
        if not any_big:
            # guard: refuse the rich path if the file on disk actually has big sheets
            if os.path.exists(file) and any(_is_big(nr, nc) for _, nr, nc in _xlsx_dims(file)):
                raise ValueError("workbook has large sheets; reload before saving")
            _build_workbook(sheets_payload).save(tmp)
        else:
            meta = _ensure_cache(file)
            _persist_edits(file, sheets_payload, meta)
            _write_xlsx_streamed(tmp, file, sheets_payload, meta)
        os.replace(tmp, file)
        if any_big:
            _rekey_cache(file)
    elif ext == ".csv":
        sp = sheets_payload[0]
        if sp.get("big"):
            meta = _ensure_cache(file)
            _, sh = _sheet_meta(meta, sp["name"])
            _persist_edits(file, [sp], meta)
            con = _duck()
            sel = ", ".join(f"c{i}" for i in range(sh["ncols"]))
            con.execute(
                f"COPY (SELECT {sel} FROM read_parquet({_q(_sheet_parquet(file, sh))}) ORDER BY __rid) "
                f"TO {_q(tmp)} (FORMAT CSV, HEADER false)"
            )
        else:
            import csv

            with open(tmp, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                for row in sp["rows"]:
                    w.writerow(row)
        os.replace(tmp, file)
        if sp.get("big"):
            _rekey_cache(file)
    elif ext == ".parquet":
        sp = sheets_payload[0]
        meta = _ensure_cache(file)
        _, sh = _sheet_meta(meta, sp["name"])
        names = sh.get("header") or [f"col{i}" for i in range(sh["ncols"])]
        con = _duck()
        if sp.get("big"):
            _persist_edits(file, [sp], meta)
            con.execute(f"CREATE TABLE t AS SELECT * FROM read_parquet({_q(_sheet_parquet(file, sh))})")
        else:
            # small parquet edited inline: rebuild from the sent rows
            rows = sp["rows"]
            con.execute("CREATE TABLE t (__rid BIGINT, " + ", ".join(f"c{i} VARCHAR" for i in range(len(names))) + ")")
            con.executemany(
                "INSERT INTO t VALUES (" + ", ".join("?" * (len(names) + 1)) + ")",
                [[r] + [str(row[c]) if c < len(row) and str(row[c]) != "" else None for c in range(len(names))]
                 for r, row in enumerate(rows)],
            )
        sel = ", ".join(f'c{i} AS "{str(names[i]).replace(chr(34), "")}"' for i in range(len(names)))
        con.execute(f"COPY (SELECT {sel} FROM t ORDER BY __rid) TO {_q(tmp)} (FORMAT PARQUET)")
        os.replace(tmp, file)
        _rekey_cache(file)
    else:
        raise ValueError(f"cannot save file type {ext!r}")
    return {"ok": True, "mtime": os.path.getmtime(file)}


def _save_as(src, directory, name, sheets_payload):
    """Write the workbook to a user-chosen directory + name and return the new path."""
    ext = os.path.splitext(src)[1].lower() if src else ".xlsx"
    if ext not in SPREADSHEET_EXTS:
        ext = ".xlsx"
    name = _safe_name(name, "workbook")
    if not name.lower().endswith(ext):
        name += ext
    dest = os.path.join(os.path.abspath(os.path.expanduser(directory or "~")), name)
    src = os.path.abspath(src) if src else ""
    if src and os.path.exists(src) and src != dest:
        import shutil

        shutil.copyfile(src, dest)  # big sheets save via the copy's rebuilt cache
    out = _save(dest, sheets_payload, "")
    out["file"] = dest.replace(os.sep, "/")
    return out


# ---------- file browsing (Save as… folder picker) ----------

def _listdir(path):
    path = os.path.abspath(os.path.expanduser(path or "~"))
    if not os.path.isdir(path):
        path = os.path.dirname(path) or "/"
    # forward slashes on every platform: the browser's crumb/join logic is "/"-based
    parent = (os.path.dirname(path) or path).replace(os.sep, "/")  # dirname(root) == root
    path = path.replace(os.sep, "/")
    dirs, files = [], []
    try:
        names = os.listdir(path)
    except OSError as e:
        return {"error": str(e), "path": path, "parent": parent, "dirs": [], "files": []}
    for name in names:
        if name.startswith("."):
            continue
        full = os.path.join(path, name)
        try:
            if os.path.isdir(full):
                dirs.append(name)
            elif name.lower().endswith(SPREADSHEET_EXTS):
                files.append({"name": name, "size": os.path.getsize(full)})
        except OSError:
            continue
    dirs.sort(key=str.lower)
    files.sort(key=lambda f: f["name"].lower())
    return {"path": path, "parent": parent, "dirs": dirs, "files": files}


# ---------- export ----------

def _latin1(s):
    return str(s).encode("latin-1", "replace").decode("latin-1")


def _pdf_from_rows(rows, title, dest):
    from fpdf import FPDF

    n_cols = max((len(r) for r in rows), default=1)
    pdf = FPDF(orientation="L" if n_cols > 7 else "P", unit="mm", format="A4")
    pdf.set_auto_page_break(True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, _latin1(title))
    pdf.ln(11)
    pdf.set_font("Helvetica", size=7 if n_cols > 12 else 8)
    with pdf.table(first_row_as_headings=True) as table:
        for row in rows:
            trow = table.row()
            for i in range(n_cols):
                trow.cell(_latin1(row[i] if i < len(row) else ""))
    pdf.output(dest)


def _export(kind, name, payload):
    os.makedirs(EXPORTS, exist_ok=True)
    base = _safe_name(name, "export")
    base = re.sub(r"\.(csv|xlsx|pdf|parquet)$", "", base, flags=re.I)

    if payload.get("big"):
        file = payload["file"]
        meta = _ensure_cache(file)
        _, sh = _sheet_meta(meta, payload["sheet"])
        con = _duck()
        _table_with_edits(con, _sheet_parquet(file, sh), payload.get("edits"), sh["ncols"])
        sel = ", ".join(f"c{i}" for i in range(sh["ncols"]))
        if kind == "csv":
            dest = os.path.join(EXPORTS, base + ".csv")
            con.execute(f"COPY (SELECT {sel} FROM t ORDER BY __rid) TO {_q(dest)} (FORMAT CSV, HEADER false)")
        elif kind == "parquet":
            names = sh.get("header") or [f"col{i}" for i in range(sh["ncols"])]
            sel2 = ", ".join(f'c{i} AS "{str(names[i]).replace(chr(34), "")}"' for i in range(sh["ncols"]))
            dest = os.path.join(EXPORTS, base + ".parquet")
            con.execute(f"COPY (SELECT {sel2} FROM t ORDER BY __rid) TO {_q(dest)} (FORMAT PARQUET)")
        elif kind == "xlsx":
            if sh["nrows"] > XLSX_SAVE_MAX_ROWS:
                raise ValueError(f"{sh['nrows']:,} rows is too large for xlsx export — use CSV or Parquet.")
            import openpyxl

            dest = os.path.join(EXPORTS, base + ".xlsx")
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(str(sh["name"])[:31])
            cur = con.execute(f"SELECT {sel} FROM t ORDER BY __rid")
            while True:
                chunk = cur.fetchmany(10_000)
                if not chunk:
                    break
                for row in chunk:
                    ws.append([_cell_in(v) if v is not None else None for v in row])
            wb.save(dest)
        elif kind == "pdf":
            cap = 2000
            rows = con.execute(f"SELECT {sel} FROM t ORDER BY __rid LIMIT {cap}").fetchall()
            rows = [["" if v is None else v for v in r] for r in rows]
            if sh["nrows"] > cap:
                rows.append([f"… {sh['nrows'] - cap:,} more rows not shown (PDF capped at {cap:,})"])
            dest = os.path.join(EXPORTS, base + ".pdf")
            _pdf_from_rows(rows, payload.get("title") or base, dest)
        else:
            raise ValueError(f"unknown export kind {kind!r}")
    elif kind == "csv":
        import csv

        dest = os.path.join(EXPORTS, base + ".csv")
        with open(dest, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for row in payload["rows"]:
                w.writerow(row)
    elif kind == "parquet":
        con = _duck()
        rows = payload["rows"]
        nc = max((len(r) for r in rows), default=1)
        con.execute("CREATE TABLE t (" + ", ".join(f'"col{i}" VARCHAR' for i in range(nc)) + ")")
        con.executemany(
            "INSERT INTO t VALUES (" + ", ".join("?" * nc) + ")",
            [[str(row[c]) if c < len(row) and str(row[c]) != "" else None for c in range(nc)] for row in rows],
        )
        dest = os.path.join(EXPORTS, base + ".parquet")
        con.execute(f"COPY t TO {_q(dest)} (FORMAT PARQUET)")
    elif kind == "xlsx":
        dest = os.path.join(EXPORTS, base + ".xlsx")
        _build_workbook(payload["sheets"]).save(dest)
    elif kind == "pdf":
        dest = os.path.join(EXPORTS, base + ".pdf")
        _pdf_from_rows(payload["rows"], payload.get("title") or base, dest)
    else:
        raise ValueError(f"unknown export kind {kind!r}")
    return {"file": dest, "name": os.path.basename(dest)}


# ---------- scripting ----------

def _run_script(script, payload):
    import contextlib
    import io

    sheets = payload["sheets"]
    env = {"sheets": sheets}
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        exec(script, env)  # local trusted code, same standing as any .py in the project
    out = env["sheets"]
    clean = []
    for sh in out:
        rows = [[_cell_out(v) for v in row] for row in sh["rows"]]
        clean.append({"name": str(sh["name"]), "rows": rows})
    return {"sheets": clean, "output": buf.getvalue()}


# Bare main() (no @fused.udf): the builtin executor calls main() directly —
# a udf wrapper hides the signature and hangs on hosted auth. The fused
# engine's compat bridge accepts a bare main() too, so this runs under both.
def main(
    action: str = "load",
    file: str = "",
    sheet: str = "",
    data: str = "",
    expected_mtime: str = "",
    name: str = "",
    directory: str = "",
    kind: str = "",
    script: str = "",
    offset: int = 0,
    limit: int = 1000,
    sort: str = "",
    filters: str = "",
    col: int = -1,
):
    if action == "load":
        if not file:
            raise ValueError("no file given")
        return _load(file)
    if action == "rows":
        meta = _ensure_cache(file)
        _, sh = _sheet_meta(meta, sheet)
        return _query_rows(
            _sheet_parquet(file, sh), sh["ncols"], offset, min(limit, 10_000),
            json.loads(sort) if sort else None,
            json.loads(filters) if filters else None,
        )
    if action == "distinct":
        meta = _ensure_cache(file)
        _, sh = _sheet_meta(meta, sheet)
        return _query_distinct(
            _sheet_parquet(file, sh), int(col),
            json.loads(filters) if filters else None,
        )
    if action == "save":
        return _save(file, json.loads(data), expected_mtime)
    if action == "save_as":
        return _save_as(file, directory, name, json.loads(data))
    if action == "listdir":
        return _listdir(file)
    if action == "export":
        return _export(kind, name, json.loads(data))
    if action == "run_script":
        return _run_script(script, json.loads(data))
    raise ValueError(f"unknown action {action!r}")
