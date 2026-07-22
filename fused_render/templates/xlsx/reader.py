"""Reader backing xlsx/template.html. Returns a JSON-safe page of one sheet.

openpyxl in read_only mode streams rows without loading the whole workbook into
memory, so we iterate once: first row is the header, the rest are data. Only the
requested page is collected; total_rows is the honest data-row count.
"""

import datetime
import decimal

import openpyxl


def _jsonify(value):
    """Coerce an openpyxl cell value into something json.dumps can encode."""
    if value is None:
        return None
    if isinstance(value, (bytes, bytearray)):
        return value.hex()
    if isinstance(value, decimal.Decimal):
        return str(value)
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    return value


def main(file: str, sheet: str = "", offset: int = 0, limit: int = 100) -> dict:
    # data_only=True returns the last-computed value of formula cells rather than
    # the formula string — matches what a spreadsheet viewer shows.
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    try:
        sheets = wb.sheetnames
        active = sheet if sheet in sheets else (sheets[0] if sheets else "")
        ws = wb[active] if active else None

        columns = []
        rows = []
        total_rows = 0
        if ws is not None:
            for i, raw in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    # First row is the header; blank cells get positional names
                    # so every column is addressable.
                    columns = [str(v) if v is not None else f"col{j}" for j, v in enumerate(raw)]
                    continue
                data_idx = i - 1  # 0-based index among data rows
                total_rows += 1
                if offset <= data_idx < offset + limit:
                    rows.append(
                        {
                            columns[j] if j < len(columns) else f"col{j}": _jsonify(v)
                            for j, v in enumerate(raw)
                        }
                    )
        return {
            "sheets": sheets,
            "sheet": active,
            "columns": columns,
            "rows": rows,
            "total_rows": total_rows,
        }
    finally:
        wb.close()
